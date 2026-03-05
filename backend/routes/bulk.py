"""
Bulk XLSX ingestion routes.
"""
from io import BytesIO
from typing import Any, Optional
import re
import zipfile
import xml.etree.ElementTree as ET

from fastapi import APIRouter, HTTPException, Query, Request
from pydantic import BaseModel, Field

from job_service import create_optimization_job
from scrapling_core.url_policy import validate_target_url

router = APIRouter()

_ALLOWED_GOALS = {"leads", "awareness", "product_info"}
_MAX_ROWS = 2000
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_REF_RE = re.compile(r"^([A-Z]+)")


class RejectedRow(BaseModel):
    row: int
    reason: str
    raw_url: Optional[str] = None


class BulkUploadResponse(BaseModel):
    submitted_count: int
    rejected_count: int
    submitted_job_ids: list[int] = Field(default_factory=list)
    rejected_rows: list[RejectedRow] = Field(default_factory=list)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _col_letters_to_index(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _xlsx_rows_minimal(content: bytes) -> list[tuple]:
    """
    Lightweight XLSX parser for simple spreadsheet ingestion.
    Handles shared strings, inline strings, and numeric/text cells.
    """
    with zipfile.ZipFile(BytesIO(content)) as zf:
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

        sheet = workbook_xml.find(f".//{{{_NS_MAIN}}}sheets/{{{_NS_MAIN}}}sheet")
        if sheet is None:
            return []
        rel_id = sheet.attrib.get(f"{{{_NS_DOC_REL}}}id")
        if not rel_id:
            return []

        target = None
        for rel in rels_xml.findall(f".//{{{_NS_PKG_REL}}}Relationship"):
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib.get("Target")
                break
        if not target:
            return []

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            shared_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in shared_root.findall(f".//{{{_NS_MAIN}}}si"):
                parts = []
                for t in si.findall(f".//{{{_NS_MAIN}}}t"):
                    parts.append(t.text or "")
                shared_strings.append("".join(parts))

        normalized_target = target.replace("\\", "/")
        while normalized_target.startswith("../"):
            normalized_target = normalized_target[3:]
        sheet_path = (
            normalized_target
            if normalized_target.startswith("xl/")
            else f"xl/{normalized_target.lstrip('/')}"
        )
        sheet_root = ET.fromstring(zf.read(sheet_path))

        rows_out: list[tuple] = []
        for row in sheet_root.findall(f".//{{{_NS_MAIN}}}sheetData/{{{_NS_MAIN}}}row"):
            row_map: dict[int, Any] = {}
            for cell in row.findall(f"{{{_NS_MAIN}}}c"):
                ref = cell.attrib.get("r", "")
                match = _CELL_REF_RE.match(ref)
                if not match:
                    continue
                col_idx = _col_letters_to_index(match.group(1))
                cell_type = cell.attrib.get("t", "")
                value = ""
                if cell_type == "inlineStr":
                    t = cell.find(f".//{{{_NS_MAIN}}}t")
                    value = t.text if t is not None else ""
                else:
                    v = cell.find(f"{{{_NS_MAIN}}}v")
                    raw = v.text if v is not None else ""
                    if cell_type == "s":
                        try:
                            ss_idx = int(raw)
                            value = shared_strings[ss_idx] if 0 <= ss_idx < len(shared_strings) else ""
                        except Exception:
                            value = ""
                    else:
                        value = raw
                row_map[col_idx] = value

            if not row_map:
                rows_out.append(tuple())
                continue

            width = max(row_map.keys()) + 1
            rows_out.append(tuple(row_map.get(i) for i in range(width)))

        return rows_out


@router.post("/bulk/upload", response_model=BulkUploadResponse)
async def bulk_upload(
    request: Request,
    filename: str = Query(default="upload.xlsx"),
):
    safe_name = (filename or "").lower()
    if not safe_name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        content = await request.body()
        if not content:
            raise ValueError("Upload payload is empty.")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            parsed_rows = list(wb.active.iter_rows(values_only=True))
        except Exception:
            parsed_rows = _xlsx_rows_minimal(content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {exc}") from exc

    rows = iter(parsed_rows)
    header = next(rows, None)
    if not header:
        raise HTTPException(status_code=400, detail="XLSX is empty.")

    headers = [_normalize_header(h) for h in header]
    if "url" not in headers:
        raise HTTPException(status_code=400, detail="XLSX must include a 'url' column.")

    idx_url = headers.index("url")
    idx_keyword = headers.index("keyword") if "keyword" in headers else None
    idx_goal = headers.index("goal") if "goal" in headers else None
    idx_num = headers.index("num_competitors") if "num_competitors" in headers else None

    submitted_ids: list[int] = []
    rejected: list[RejectedRow] = []
    processed = 0

    for row_number, row in enumerate(rows, start=2):
        if processed >= _MAX_ROWS:
            rejected.append(
                RejectedRow(
                    row=row_number,
                    reason=f"Row limit exceeded (max {_MAX_ROWS}).",
                    raw_url=str(row[idx_url]) if idx_url < len(row) and row[idx_url] else None,
                )
            )
            continue

        raw_url = row[idx_url] if idx_url < len(row) else None
        if not raw_url:
            continue
        url = str(raw_url).strip()
        if not url:
            continue

        keyword = ""
        if idx_keyword is not None and idx_keyword < len(row) and row[idx_keyword] is not None:
            keyword = str(row[idx_keyword]).strip()

        goal = "leads"
        if idx_goal is not None and idx_goal < len(row) and row[idx_goal] is not None:
            candidate_goal = str(row[idx_goal]).strip().lower()
            if candidate_goal:
                goal = candidate_goal

        num_competitors = 10
        if idx_num is not None and idx_num < len(row) and row[idx_num] is not None:
            try:
                num_competitors = int(row[idx_num])
            except Exception:
                rejected.append(
                    RejectedRow(row=row_number, reason="num_competitors must be an integer.", raw_url=url)
                )
                continue

        if goal not in _ALLOWED_GOALS:
            rejected.append(
                RejectedRow(
                    row=row_number,
                    reason=f"goal must be one of: {', '.join(sorted(_ALLOWED_GOALS))}.",
                    raw_url=url,
                )
            )
            continue
        if num_competitors < 3 or num_competitors > 20:
            rejected.append(
                RejectedRow(
                    row=row_number,
                    reason="num_competitors must be between 3 and 20.",
                    raw_url=url,
                )
            )
            continue

        try:
            validate_target_url(url)
        except ValueError as exc:
            rejected.append(RejectedRow(row=row_number, reason=str(exc), raw_url=url))
            continue

        try:
            job = create_optimization_job(
                url=url,
                keyword=keyword,
                goal=goal,
                num_competitors=num_competitors,
                pipeline_mode="full",
                schedule_id=None,
            )
            submitted_ids.append(job.id)
            processed += 1
        except Exception as exc:
            rejected.append(
                RejectedRow(
                    row=row_number,
                    reason=f"Failed to enqueue job: {exc}",
                    raw_url=url,
                )
            )

    return BulkUploadResponse(
        submitted_count=len(submitted_ids),
        rejected_count=len(rejected),
        submitted_job_ids=submitted_ids,
        rejected_rows=rejected,
    )
